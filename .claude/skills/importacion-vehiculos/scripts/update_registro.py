# -*- coding: utf-8 -*-
"""
update_registro.py — Anade una fila nueva al Registro historico de
operaciones (archivo separado de la plantilla por coche, ver tarea del
usuario: "Registro de operaciones deberia estar en otro archivo distinto").

Si el archivo de registro indicado no existe todavia, se crea una copia del
master bundled en assets/Registro_Operaciones_master.xlsx antes de anadir la
fila.

Uso:
    python update_registro.py registro.json /ruta/Registro_Operaciones_Importacion.xlsx

El JSON de entrada:
{
  "fecha_compra": "15/07/2026",
  "vehiculo": "BMW Serie 1 118i M Sport",
  "precio_compra": 21500,
  "coste_total": 23316,
  "precio_venta": null,
  "fecha_venta": null,
  "estado": "En curso",
  "notas": "Importado desde Alemania, ver Drive"
}

Nota importante sobre Google Drive: este script solo mantiene el archivo
LOCAL. El conector de Drive disponible en esta skill (create_file,
search_files, ...) no tiene una operacion de "actualizar archivo existente"
— solo puede crear archivos nuevos. Por eso el Registro no se puede
mantener "en vivo" dentro de un Google Sheet ya existente sin crear
duplicados cada vez. Recomendacion practica: mantener este .xlsx como
fichero local de referencia (idealmente en una carpeta que persista entre
sesiones) y subir una copia fresca a Drive de vez en cuando si se quiere
tener una version compartida, avisando siempre al usuario de que esa subida
crea un archivo nuevo, no actualiza el anterior.
"""
import json
import shutil
import sys
from pathlib import Path

import openpyxl

SHEET = "Registro de operaciones"
FIRST_ROW = 5
# Fila que contiene el total (no se debe pisar); se recalcula si hay que insertar filas.
TOTAL_LABEL = "Honorarios totales cobrados (operaciones cerradas)"

# Modelo de negocio: servicio de importacion por encargo, no reventa. El
# coche se matricula a nombre del cliente; "precio_venta" es en realidad el
# precio TODO INCLUIDO cobrado al cliente (coche + importacion + honorarios),
# y la columna G calcula los honorarios como precio_venta - coste_total.
COLUMNS = {
    "fecha_compra": 1,
    "vehiculo": 2,
    "precio_compra": 3,
    "coste_total": 4,
    "precio_venta": 5,       # precio todo incluido cobrado al cliente
    "fecha_venta": 6,        # fecha de cierre de la operacion
    # columna 7 (G) y 8 (H) son formulas de honorarios / honorarios-sobre-coste, no se tocan
    "estado": 9,             # "En curso" o "Cerrada"
    "notas": 10,
}


def find_total_row(ws):
    for row in range(1, ws.max_row + 5):
        if ws.cell(row=row, column=2).value == TOTAL_LABEL:
            return row
    return None


def find_empty_row(ws, total_row):
    row = FIRST_ROW
    while row < total_row:
        if not ws.cell(row=row, column=2).value:  # columna B = Vehiculo
            return row
        row += 1
    return None


def ensure_formulas(ws, row):
    if not ws.cell(row=row, column=7).value:
        ws.cell(row=row, column=7, value=f'=IF(E{row}="","",E{row}-D{row})')
    if not ws.cell(row=row, column=8).value:
        ws.cell(row=row, column=8, value=f'=IF(OR(E{row}="",D{row}=0),"",(E{row}-D{row})/D{row})')


def insert_row_before_total(ws, total_row):
    ws.insert_rows(total_row)
    new_row = total_row
    ensure_formulas(ws, new_row)
    new_total_row = total_row + 1
    return new_row, new_total_row


def append_row(ws, data, row):
    for key, col in COLUMNS.items():
        val = data.get(key)
        if val not in (None, ""):
            ws.cell(row=row, column=col, value=val)
    if not ws.cell(row=row, column=9).value:
        ws.cell(row=row, column=9, value="En curso")
    ensure_formulas(ws, row)


def main():
    if len(sys.argv) < 3:
        print("Uso: python update_registro.py registro.json /ruta/Registro_Operaciones_Importacion.xlsx [master.xlsx]")
        sys.exit(1)

    data_path = Path(sys.argv[1])
    registro_path = Path(sys.argv[2])
    master_path = Path(sys.argv[3]) if len(sys.argv) > 3 else (
        Path(__file__).resolve().parent.parent / "assets" / "Registro_Operaciones_master.xlsx"
    )

    data = json.loads(data_path.read_text(encoding="utf-8"))

    created_new = False
    if not registro_path.exists():
        registro_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(master_path, registro_path)
        created_new = True

    wb = openpyxl.load_workbook(registro_path)
    ws = wb[SHEET]

    total_row = find_total_row(ws)
    if total_row is None:
        raise RuntimeError("No se encontro la fila de total en el Registro; revisa el archivo.")

    row = find_empty_row(ws, total_row)
    if row is None:
        row, total_row = insert_row_before_total(ws, total_row)
        # Actualiza el rango del SUMIF del total para incluir la fila nueva
        total_cell = ws.cell(row=total_row, column=7)
        if total_cell.value and str(total_cell.value).startswith("=SUMIF"):
            total_cell.value = f'=SUMIF(I{FIRST_ROW}:I{row},"Vendido",G{FIRST_ROW}:G{row})'

    append_row(ws, data, row)
    wb.save(registro_path)

    print(f"OK: fila {row} escrita en {registro_path} (archivo nuevo: {created_new})")
    print("Recuerda recalcular con recalc.py (skill xlsx) antes de darlo por definitivo.")


if __name__ == "__main__":
    main()
