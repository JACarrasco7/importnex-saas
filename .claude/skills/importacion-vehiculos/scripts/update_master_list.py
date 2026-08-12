# -*- coding: utf-8 -*-
"""
update_master_list.py — Anade una fila nueva a una de las dos listas
maestras (archivos separados, no la plantilla por coche ni la ficha por
cliente): Clientes_master.xlsx o Inventario_Coches_Oferta_master.xlsx.

Uso:
    python update_master_list.py clientes datos.json /ruta/Clientes.xlsx [master.xlsx]
    python update_master_list.py inventario datos.json /ruta/Inventario_Coches_Oferta.xlsx [master.xlsx]

JSON para "clientes":
{
  "fecha_alta": "23/07/2026", "nombre": "...", "contacto": "tel / email",
  "como_llego": "Referido", "que_busca": "SUV compacto diesel",
  "presupuesto_min": 15000, "presupuesto_max": 22000, "uso": "Reventa",
  "plazo_deseado": "2 meses", "estado": "Nuevo lead",
  "enlace_ficha": "https://drive.google.com/...", "coche_asignado": ""
}

JSON para "inventario":
{
  "fecha_alta": "23/07/2026", "marca_modelo": "Audi A3 Sportback 2.0 TDI",
  "anio": 2020, "km": 68000, "coste_estimado": 16500, "precio_venta_objetivo": 19900,
  "estado": "Buscando comprador", "cliente_interesado": "",
  "enlace_ficha": "https://drive.google.com/...", "notas": "..."
}

Si el archivo de destino no existe todavia, se crea una copia del master
bundled correspondiente antes de anadir la fila. Igual que con el Registro
de operaciones, si el destino vive en Drive el protocolo es: descargar la
version mas reciente, anadir la fila sobre esa copia, recalcular y volver a
subir (ver references/google_drive.md) — este script solo edita el .xlsx que
se le indique como ruta, no habla con Drive directamente.
"""
import json
import shutil
import sys
from pathlib import Path

import openpyxl

CONFIGS = {
    "clientes": {
        "sheet": "Clientes",
        "first_row": 5,
        "key_col": 2,  # Nombre: define si una fila esta vacia
        "columns": {
            "fecha_alta": 1, "nombre": 2, "contacto": 3, "como_llego": 4,
            "que_busca": 5, "presupuesto_min": 6, "presupuesto_max": 7,
            "uso": 8, "plazo_deseado": 9, "estado": 10,
            "enlace_ficha": 11, "coche_asignado": 12,
        },
        "estado_col": 10,
        "estado_default": "Nuevo lead",
        "total_label": "Clientes activos (no cerrados ni descartados)",
        "total_col": 10,
        "total_formula": '=COUNTIFS(J{first}:J{last},"<>Cerrado - vendido",J{first}:J{last},"<>Descartado",J{first}:J{last},"<>")',
        "master_asset": "Clientes_master.xlsx",
    },
    "inventario": {
        "sheet": "Inventario",
        "first_row": 5,
        "key_col": 2,  # Marca y modelo
        "columns": {
            "fecha_alta": 1, "marca_modelo": 2, "anio": 3, "km": 4,
            "coste_estimado": 5, "precio_venta_objetivo": 6, "estado": 7,
            "cliente_interesado": 8, "enlace_ficha": 9, "notas": 10,
        },
        "estado_col": 7,
        "estado_default": "Buscando comprador",
        "total_label": "Coches disponibles ahora mismo",
        "total_col": 7,
        "total_formula": '=COUNTIF(G{first}:G{last},"Buscando comprador")',
        "master_asset": "Inventario_Coches_Oferta_master.xlsx",
    },
}


def find_total_row(ws, total_label):
    for row in range(1, ws.max_row + 5):
        if ws.cell(row=row, column=2).value == total_label:
            return row
    return None


def find_empty_row(ws, first_row, key_col, total_row):
    row = first_row
    while row < total_row:
        if not ws.cell(row=row, column=key_col).value:
            return row
        row += 1
    return None


def append_row(ws, cfg, data, row):
    for key, col in cfg["columns"].items():
        val = data.get(key)
        if val not in (None, ""):
            ws.cell(row=row, column=col, value=val)
    estado_col = cfg["estado_col"]
    if not ws.cell(row=row, column=estado_col).value:
        ws.cell(row=row, column=estado_col, value=cfg["estado_default"])


def insert_row_before_total(ws, cfg, total_row, first_row):
    ws.insert_rows(total_row)
    new_row = total_row
    new_total_row = total_row + 1
    total_cell = ws.cell(row=new_total_row, column=cfg["total_col"])
    total_cell.value = cfg["total_formula"].format(first=first_row, last=new_row)
    return new_row, new_total_row


def main():
    if len(sys.argv) < 4:
        print("Uso: python update_master_list.py <clientes|inventario> datos.json /ruta/archivo.xlsx [master.xlsx]")
        sys.exit(1)

    tipo = sys.argv[1]
    if tipo not in CONFIGS:
        print(f"Tipo desconocido: {tipo!r}. Usa 'clientes' o 'inventario'.")
        sys.exit(1)
    cfg = CONFIGS[tipo]

    data_path = Path(sys.argv[2])
    target_path = Path(sys.argv[3])
    master_path = Path(sys.argv[4]) if len(sys.argv) > 4 else (
        Path(__file__).resolve().parent.parent / "assets" / cfg["master_asset"]
    )

    data = json.loads(data_path.read_text(encoding="utf-8"))

    created_new = False
    if not target_path.exists():
        target_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(master_path, target_path)
        created_new = True

    wb = openpyxl.load_workbook(target_path)
    ws = wb[cfg["sheet"]]

    total_row = find_total_row(ws, cfg["total_label"])
    if total_row is None:
        raise RuntimeError(f"No se encontro la fila de total ('{cfg['total_label']}') en {target_path}.")

    row = find_empty_row(ws, cfg["first_row"], cfg["key_col"], total_row)
    if row is None:
        row, total_row = insert_row_before_total(ws, cfg, total_row, cfg["first_row"])

    append_row(ws, cfg, data, row)

    # IMPORTANTE (ver nota sobre sincronizacion de archivos): guardar siempre
    # primero en una ruta nueva/temporal y luego copiar sobre el destino
    # final, nunca sobrescribir en el sitio si el destino esta en una carpeta
    # montada compartida — evita quedarse con una version a medio escribir.
    tmp_path = target_path.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)
    shutil.copy(tmp_path, target_path)
    try:
        tmp_path.unlink()
    except OSError:
        # Ver nota equivalente en sync_web_data.py: en algunos entornos
        # montados borrar el .tmp intermedio da "Operation not permitted"
        # aunque la copia al destino ya se hizo bien. No es un fallo real.
        pass

    print(f"OK: fila {row} escrita en {target_path} (archivo nuevo: {created_new})")
    print("Recuerda recalcular antes de darlo por definitivo.")


if __name__ == "__main__":
    main()
