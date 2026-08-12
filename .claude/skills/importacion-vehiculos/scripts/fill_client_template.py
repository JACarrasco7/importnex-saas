# -*- coding: utf-8 -*-
"""
fill_client_template.py — Rellena una copia de la plantilla maestra de ficha
de cliente (assets/Ficha_Cliente_master.xlsx) con los datos de un cliente
concreto y, opcionalmente, la lista de coches que se le han propuesto.

Uso:
    python fill_client_template.py datos.json /ruta/salida/Ficha_Cliente_XXX.xlsx [ruta_master.xlsx]

El JSON de entrada (todos los campos opcionales, rellena solo lo que tengas):

{
  "cliente": {
    "nombre": "...", "telefono": "...", "email": "...",
    "como_llego": "Referido / RRSS / Portal / Otro",
    "fecha_alta": "23/07/2026",
    "estado": "Nuevo lead"
  },
  "busca": {
    "marca_modelo_tipo": "...", "presupuesto_min": 15000, "presupuesto_max": 22000,
    "uso_previsto": "Reventa", "plazo_deseado": "2 meses",
    "km_maximo": 100000, "anio_minimo": 2018,
    "combustible_cambio": "...", "otras_preferencias": "..."
  },
  "coches_propuestos": [
    {"fecha": "23/07/2026", "coche": "Audi A3 Sportback 2.0 TDI", "precio": 19900,
     "estado": "Propuesto", "enlace": "https://drive.google.com/...", "notas": "..."}
  ],
  "historial_contacto": [
    {"fecha": "23/07/2026", "canal": "WhatsApp", "resumen": "...", "proximos_pasos": "..."}
  ]
}

Este script solo toca 'Cliente y resumen', 'Coches propuestos' e 'Historial
de contacto'. La hoja 'Plantillas de mensaje' (copiada de la plantilla de
coche) no se toca, es reutilizable tal cual.

IMPORTANTE: tras ejecutar este script hay que recalcular el libro con
LibreOffice antes de entregarlo (openpyxl no calcula formulas). En este
entorno, si recalc.py del skill 'xlsx' da timeout, usar directamente:
    soffice --headless --norestore --convert-to xlsx --outdir <dir> archivo.xlsx
y sustituir el archivo original por el convertido.
"""
import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

FONT_NAME = "Arial"
BLUE_INPUT = "0000FF"
DARK_TEXT = "22262B"
YELLOW_FILL = "FFF2AC"

fill_yellow = PatternFill("solid", fgColor=YELLOW_FILL)


def f_input():
    return Font(name=FONT_NAME, size=10, color=BLUE_INPUT)


CLIENTE_SHEET = "Cliente y resumen"
CLIENTE_ROW_MAP = {
    "nombre": 6,
    "telefono": 7,
    "email": 8,
    "como_llego": 9,
    "fecha_alta": 10,
    "estado": 11,
}

BUSCA_ROW_MAP = {
    "marca_modelo_tipo": 14,
    "presupuesto_min": 15,
    "presupuesto_max": 16,
    "uso_previsto": 17,
    "plazo_deseado": 18,
    "km_maximo": 19,
    "anio_minimo": 20,
    "combustible_cambio": 21,
    "otras_preferencias": 22,
}

ROW_COCHE_ELEGIDO = 27
ROW_ENLACE_OPERACION = 28

PROPUESTOS_SHEET = "Coches propuestos"
PROPUESTOS_FIRST_ROW = 5
PROPUESTOS_LAST_ROW = 16  # 12 filas disponibles en el master; se puede ampliar a mano si hace falta

HISTORIAL_SHEET = "Historial de contacto"
HISTORIAL_FIRST_ROW = 5
HISTORIAL_LAST_ROW = 19  # 15 filas disponibles en el master


def fill_cliente(ws, cliente: dict):
    for key, row in CLIENTE_ROW_MAP.items():
        if key not in cliente:
            continue
        val = cliente.get(key)
        cell = ws.cell(row=row, column=2, value=val if val not in ("", None) else None)
        cell.font = f_input()


def fill_busca(ws, busca: dict):
    for key, row in BUSCA_ROW_MAP.items():
        if key not in busca:
            continue
        val = busca.get(key)
        cell = ws.cell(row=row, column=2, value=val if val not in ("", None) else None)
        cell.font = f_input()


def fill_coche_elegido(ws, coche_elegido, enlace_operacion):
    if coche_elegido not in (None, ""):
        c = ws.cell(row=ROW_COCHE_ELEGIDO, column=2, value=coche_elegido)
        c.font = f_input(); c.fill = fill_yellow
    if enlace_operacion not in (None, ""):
        c = ws.cell(row=ROW_ENLACE_OPERACION, column=2, value=enlace_operacion)
        c.font = f_input(); c.fill = fill_yellow


def fill_propuestos(ws, propuestos: list):
    if not propuestos:
        return
    cols = ["fecha", "coche", "precio", "estado", "enlace", "notas"]
    row = PROPUESTOS_FIRST_ROW
    for entry in propuestos:
        if row > PROPUESTOS_LAST_ROW:
            break
        for ci, key in enumerate(cols):
            val = entry.get(key)
            cell = ws.cell(row=row, column=2 + ci, value=val if val not in ("", None) else None)
            cell.font = f_input()
        row += 1


def fill_historial(ws, historial: list):
    if not historial:
        return
    cols = ["fecha", "canal", "resumen", "proximos_pasos"]
    row = HISTORIAL_FIRST_ROW
    for entry in historial:
        if row > HISTORIAL_LAST_ROW:
            break
        for ci, key in enumerate(cols):
            val = entry.get(key)
            cell = ws.cell(row=row, column=2 + ci, value=val if val not in ("", None) else None)
            cell.font = f_input()
        row += 1


def main():
    if len(sys.argv) < 3:
        print("Uso: python fill_client_template.py datos.json /ruta/salida.xlsx [ruta_master.xlsx]")
        sys.exit(1)

    data_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2])
    master_path = Path(sys.argv[3]) if len(sys.argv) > 3 else (
        Path(__file__).resolve().parent.parent / "assets" / "Ficha_Cliente_master.xlsx"
    )

    data = json.loads(data_path.read_text(encoding="utf-8"))

    wb = openpyxl.load_workbook(master_path)
    fill_cliente(wb[CLIENTE_SHEET], data.get("cliente", {}))
    fill_busca(wb[CLIENTE_SHEET], data.get("busca", {}))
    fill_coche_elegido(wb[CLIENTE_SHEET], data.get("coche_elegido"), data.get("enlace_operacion"))
    fill_propuestos(wb[PROPUESTOS_SHEET], data.get("coches_propuestos", []))
    fill_historial(wb[HISTORIAL_SHEET], data.get("historial_contacto", []))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"OK: guardado en {out_path}")
    print("Recuerda recalcular antes de entregar el archivo.")


if __name__ == "__main__":
    main()
