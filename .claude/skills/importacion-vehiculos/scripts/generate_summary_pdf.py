# -*- coding: utf-8 -*-
"""
generate_summary_pdf.py — Genera un resumen ejecutivo de una pagina (PDF)
para una operacion de importacion, a partir del .xlsx ya rellenado y
recalculado (fill_template.py + recalc.py) y del mismo JSON de datos que se
uso para rellenarlo.

Uso:
    python generate_summary_pdf.py Plantilla_XXX.xlsx datos.json /ruta/salida/Resumen_XXX.pdf

Por que lee el xlsx en vez de recalcular el mismo: las cifras clave (coste
total, honorarios, precio todo incluido) son formulas que dependen de toda
la logica de 'Numeros'. Es
mas fiable leer los valores ya recalculados por LibreOffice que reimplementar
el calculo aqui por segunda vez y arriesgarse a que diverjan.
"""
import json
import sys
from pathlib import Path

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
)

NAVY = colors.HexColor("#0B1F3A")
ORANGE = colors.HexColor("#E8590C")
GREY = colors.HexColor("#5A6472")
GREEN = colors.HexColor("#1E7B34")
RED = colors.HexColor("#B3261E")
LIGHT = colors.HexColor("#F4F6F8")


def money(v):
    if v in (None, ""):
        return "-"
    try:
        return f"{float(v):,.0f} EUR".replace(",", ".")
    except (TypeError, ValueError):
        return str(v)


def pct(v):
    if v in (None, ""):
        return "-"
    try:
        return f"{float(v) * 100:.1f} %"
    except (TypeError, ValueError):
        return str(v)


def read_summary(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Vehiculo y resumen"]
    data = {
        "marca_modelo": ws["B6"].value,
        "motorizacion": ws["B7"].value,
        "anio": ws["B12"].value,
        "km": ws["B13"].value,
        "precio_anuncio": ws["B18"].value,
        "co2": ws["B19"].value,
        "vendedor": ws["B24"].value,
        "coste_min": ws["B43"].value,
        "coste_max": ws["B44"].value,
        "precio_todo_incluido_min": ws["B45"].value,
        "precio_todo_incluido_max": ws["B46"].value,
        "honorarios": ws["B47"].value,
        "honorarios_pct": ws["B48"].value,
        "precio_medio_comparables": ws["B49"].value,
        "insp_ok": ws["B52"].value,
        "insp_revisar": ws["B53"].value,
        "insp_total": ws["B54"].value,
        "doc_recibidos": ws["B55"].value,
        "doc_total": ws["B56"].value,
        "investigacion": {},
    }
    invest_labels = {
        33: "problemas_comunes",
        34: "recalls",
        35: "precio_mercado",
        36: "fiabilidad",
        37: "otros",
    }
    for row, key in invest_labels.items():
        hallazgo = ws.cell(row=row, column=2).value
        fuente = ws.cell(row=row, column=3).value
        if hallazgo:
            data["investigacion"][key] = {"hallazgo": hallazgo, "fuente": fuente}
    return data


def build_pdf(summary, extra, out_path):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCustom", parent=styles["Title"], textColor=NAVY, fontSize=18, spaceAfter=2)
    subtitle_style = ParagraphStyle("SubtitleCustom", parent=styles["Normal"], textColor=GREY, fontSize=9, spaceAfter=10)
    h2 = ParagraphStyle("H2Custom", parent=styles["Heading2"], textColor=ORANGE, fontSize=11, spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("BodyCustom", parent=styles["Normal"], fontSize=9, leading=12)
    small = ParagraphStyle("SmallCustom", parent=styles["Normal"], fontSize=7.5, textColor=GREY, leading=10)

    story = []
    marca = summary.get("marca_modelo") or "Vehiculo sin identificar"
    story.append(Paragraph(marca, title_style))
    detalle = f"{summary.get('motorizacion') or ''} - {summary.get('anio') or 's/f'} - {summary.get('km') or '?'} km"
    story.append(Paragraph(detalle, subtitle_style))
    story.append(HRFlowable(width="100%", color=NAVY, thickness=1.2))

    # --- Bandera de riesgo ---
    precio_incluido_min = summary.get("precio_todo_incluido_min")
    precio_medio_comp = summary.get("precio_medio_comparables")
    revisar = summary.get("insp_revisar") or 0
    flags = []
    if precio_incluido_min is not None and precio_medio_comp:
        try:
            if float(precio_incluido_min) > float(precio_medio_comp) * 1.05:
                flags.append("El precio todo incluido queda por encima de la media de comparables de mercado (semaforo rojo).")
        except (TypeError, ValueError):
            pass
    if revisar and revisar > 0:
        flags.append(f"{revisar} punto(s) de la inspeccion marcados como 'Revisar'.")
    if extra.get("red_flags"):
        flags.extend(extra["red_flags"])
    if flags:
        flag_style = ParagraphStyle("Flag", parent=body, textColor=RED, backColor=colors.HexColor("#F8C9C4"))
        for f in flags:
            story.append(Paragraph(f"<b>AVISO:</b> {f}", flag_style))
        story.append(Spacer(1, 4))

    # --- KPIs ---
    story.append(Paragraph("Coste, honorarios y precio al cliente", h2))
    kpi_data = [
        ["Coste total (min-max)", f"{money(summary.get('coste_min'))} - {money(summary.get('coste_max'))}"],
        ["Honorarios de servicio", money(summary.get("honorarios"))],
        ["Precio todo incluido al cliente (min-max)", f"{money(summary.get('precio_todo_incluido_min'))} - {money(summary.get('precio_todo_incluido_max'))}"],
        ["Honorarios (%) sobre coste minimo", pct(summary.get("honorarios_pct"))],
        ["Precio medio de comparables (mercado)", money(summary.get("precio_medio_comparables"))],
        ["Precio del anuncio (origen)", money(summary.get("precio_anuncio"))],
    ]
    t = Table(kpi_data, colWidths=[75 * mm, 75 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), GREY),
        ("TEXTCOLOR", (1, 0), (1, -1), NAVY),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#C7CDD4")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(t)

    # --- Progreso ---
    story.append(Paragraph("Progreso de revision", h2))
    prog = (
        f"Inspeccion: {summary.get('insp_ok') or 0} OK - {summary.get('insp_revisar') or 0} a revisar - "
        f"{summary.get('insp_total') or 0} puntos totales.  &nbsp;&nbsp; "
        f"Documentacion: {summary.get('doc_recibidos') or 0} / {summary.get('doc_total') or 0} recibidos."
    )
    story.append(Paragraph(prog, body))

    # --- Investigacion ---
    story.append(Paragraph("Investigacion del modelo", h2))
    labels = {
        "problemas_comunes": "Problemas / averias comunes",
        "recalls": "Recalls oficiales",
        "precio_mercado": "Rango de precio de mercado",
        "fiabilidad": "Fiabilidad general",
        "otros": "Otros datos relevantes",
    }
    inv = summary.get("investigacion") or {}
    for key, label in labels.items():
        entry = inv.get(key)
        if not entry:
            continue
        hallazgo = entry.get("hallazgo", "")
        fuente = entry.get("fuente", "")
        story.append(Paragraph(f"<b>{label}:</b> {hallazgo}", body))
        if fuente:
            story.append(Paragraph(f"Fuente: {fuente}", small))
        story.append(Spacer(1, 3))

    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", color=colors.HexColor("#C7CDD4"), thickness=0.5))
    story.append(Paragraph(
        "Resumen generado automaticamente a partir del anuncio y de investigacion web. "
        "Cifras orientativas: confirmar CO2 real (COC), estado del vehiculo en la inspeccion fisica "
        "y condiciones fiscales con gestoria antes de comprometerse.",
        small,
    ))

    doc = SimpleDocTemplate(
        str(out_path), pagesize=A4,
        topMargin=18 * mm, bottomMargin=15 * mm, leftMargin=18 * mm, rightMargin=18 * mm,
    )
    doc.build(story)


def main():
    if len(sys.argv) < 4:
        print("Uso: python generate_summary_pdf.py Plantilla_XXX.xlsx datos.json /ruta/salida/Resumen_XXX.pdf")
        sys.exit(1)
    xlsx_path = Path(sys.argv[1])
    json_path = Path(sys.argv[2])
    out_path = Path(sys.argv[3])

    summary = read_summary(xlsx_path)
    extra = json.loads(json_path.read_text(encoding="utf-8")) if json_path.exists() else {}

    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_pdf(summary, extra, out_path)
    print(f"OK: PDF guardado en {out_path}")


if __name__ == "__main__":
    main()
