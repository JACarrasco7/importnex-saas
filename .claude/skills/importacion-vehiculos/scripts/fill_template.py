# -*- coding: utf-8 -*-
"""
fill_template.py — Rellena una copia de la plantilla maestra de importacion
de vehiculos (assets/Plantilla_Importacion_Vehiculos_master.xlsx) con los
datos de un coche concreto.

Uso:
    python fill_template.py datos.json /ruta/salida/Plantilla_XXX.xlsx [ruta_master.xlsx]

El JSON de entrada debe tener esta forma (todos los campos son opcionales,
rellena solo lo que tengas; deja huecos "" o null si no hay dato):

{
  "vehicle": {
    "marca_modelo": "...", "motorizacion": "...", "combustible": "...",
    "cambio": "...", "traccion": "...", "puertas_plazas": "...",
    "anio_matriculacion": "...", "kilometraje": 12345, "propietarios": 2,
    "color_exterior": "...", "color_interior": "...", "equipamiento": "...",
    "precio_anuncio": 12990, "co2": 109, "garantia": "...",
    "accidentes_declarados": "...", "historial_mantenimiento": "...",
    "vendedor_tipo": "...", "vendedor_nombre_ubicacion": "...",
    "url_anuncio": "...", "enlace_fotos": "...", "contacto": "...",
    "fecha_captura": "..."
  },
  "investigacion": {
    "problemas_comunes": {"hallazgo": "...", "fuente": "..."},
    "recalls":           {"hallazgo": "...", "fuente": "..."},
    "precio_mercado":    {"hallazgo": "...", "fuente": "..."},
    "fiabilidad":        {"hallazgo": "...", "fuente": "..."},
    "otros":             {"hallazgo": "...", "fuente": "..."}
  },
  "comparables": [
    {"portal": "...", "url": "...", "motor": "...", "cv": 150, "anio": 2019,
     "km": 80000, "precio": 16000, "ubicacion": "...", "notas": "..."}
  ],
  "vendedor_contacto": {"nombre": "...", "telefono": "...", "email": "...", "notas": "..."},
  "honorarios_servicio": 1500,
  "anuncio": {
    "titulo": "...",
    "descripcion_corta": "Texto para RRSS (Instagram/Facebook/grupos)",
    "descripcion_larga": "Texto para portal de venta (Wallapop/Coches.net/AutoScout24)",
    "hashtags": "#coche #importacion #..."
  }
}

IMPORTANTE — modelo de negocio: JJ Import Motors presta un SERVICIO de
gestion de importacion (nunca reventa). El coche se matricula a nombre del
cliente; JJ Import Motors cobra unos honorarios fijos por tramo de precio
(hasta 15.000E -> 1.500E; 15.000-30.000E -> 2.000-2.500E; +30.000E ->
2.500-3.500E). `honorarios_servicio` es ese importe, NO un margen de
reventa — la plantilla calcula sola el "precio todo incluido al cliente"
como coste total + honorarios. No redactes nada (anuncios, mensajes) que
suene a "vendo este coche"; usa siempre lenguaje de servicio ("te lo
importo", "gestiono la importacion", "matriculado a tu nombre").

Este script SOLO toca las celdas de datos de la operacion (hoja 'Vehiculo y
resumen', la tabla de comparables + honorarios de servicio en 'Numeros', y
la fila del vendedor en 'Contactos clave'). El resto de la plantilla
(checklists, cronograma, plantillas de mensaje) se deja intacta tal y como
esta en el master, porque son secciones reutilizables pensadas para
cualquier operacion.

IMPORTANTE: tras ejecutar este script hay que recalcular el libro con
LibreOffice (recalc.py del skill 'xlsx') antes de entregarlo, porque
openpyxl no calcula formulas.
"""
import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

FONT_NAME = "Arial"
BLUE_INPUT = "0000FF"
DARK_TEXT = "22262B"
YELLOW_FILL = "FFF2AC"

fill_yellow = PatternFill("solid", fgColor=YELLOW_FILL)


def f_input():
    return Font(name=FONT_NAME, size=10, color=BLUE_INPUT)


def f_label():
    return Font(name=FONT_NAME, size=10, color=DARK_TEXT)


# Mapa de filas fijo de la hoja "Vehiculo y resumen" (bloque DATOS DEL VEHICULO).
# Debe coincidir exactamente con build_template_v3.py. Si la plantilla maestra
# cambia de estructura, actualizar este mapa.
VEHICLE_ROW_MAP = {
    "marca_modelo": 6,
    "motorizacion": 7,
    "combustible": 8,
    "cambio": 9,
    "traccion": 10,
    "puertas_plazas": 11,
    "anio_matriculacion": 12,
    "kilometraje": 13,
    "propietarios": 14,
    "color_exterior": 15,
    "color_interior": 16,
    "equipamiento": 17,
    "precio_anuncio": 18,
    "co2": 19,
    "garantia": 20,
    "accidentes_declarados": 21,
    "historial_mantenimiento": 22,
    "vendedor_tipo": 23,
    "vendedor_nombre_ubicacion": 24,
    "url_anuncio": 25,
    "enlace_fotos": 26,
    "contacto": 27,
    "fecha_captura": 28,
}
# Campos cuyo valor debe marcarse en amarillo (dato clave a confirmar).
ASSUMPTION_FIELDS = {"co2", "accidentes_declarados"}

# Bloque INVESTIGACION DEL MODELO: campo -> fila (columna B=Hallazgo, C=Fuente)
INVEST_ROW_MAP = {
    "problemas_comunes": 33,
    "recalls": 34,
    "precio_mercado": 35,
    "fiabilidad": 36,
    "otros": 37,
}

COMPARABLES_SHEET = "Numeros"
COMPARABLES_FIRST_ROW = 42
COMPARABLES_LAST_ROW = 50  # 9 filas disponibles en el master
HONORARIOS_ROW = 32         # "Honorarios de servicio" dentro de HONORARIOS Y PRECIO TODO INCLUIDO

CONTACTOS_SHEET = "Contactos clave"
CONTACTOS_VENDEDOR_ROW = 5  # columna B=Nombre/Entidad, C=Telefono, D=Email, E=Notas

# Hoja "Anuncio de venta" (copy listo para publicar). Columna C = valor.
ANUNCIO_SHEET = "Anuncio de venta"
ANUNCIO_ROW_TITULO = 5
ANUNCIO_ROW_DESC_CORTA = 10
ANUNCIO_ROW_DESC_LARGA = 14
ANUNCIO_ROW_HASHTAGS = 27


def fill_vehicle(ws, vehicle: dict):
    for key, row in VEHICLE_ROW_MAP.items():
        if key not in vehicle:
            continue
        val = vehicle.get(key)
        cell = ws.cell(row=row, column=2)
        cell.value = val if val not in ("", None) else None
        cell.font = f_input()
        if key in ASSUMPTION_FIELDS:
            cell.fill = fill_yellow


def fill_investigacion(ws, investigacion: dict):
    for key, row in INVEST_ROW_MAP.items():
        entry = investigacion.get(key) if investigacion else None
        if not entry:
            continue
        hallazgo = entry.get("hallazgo") if isinstance(entry, dict) else entry
        fuente = entry.get("fuente") if isinstance(entry, dict) else None
        hc = ws.cell(row=row, column=2, value=hallazgo or None)
        hc.font = f_input()
        hc.fill = fill_yellow
        if fuente:
            ws.cell(row=row, column=3, value=fuente).font = f_label()


def fill_comparables(ws, comparables: list):
    if not comparables:
        return
    cols = ["portal", "url", "motor", "cv", "anio", "km", "precio", "ubicacion", "notas"]
    row = COMPARABLES_FIRST_ROW
    for entry in comparables:
        if row > COMPARABLES_LAST_ROW:
            break
        for ci, key in enumerate(cols):
            val = entry.get(key)
            cell = ws.cell(row=row, column=1 + ci, value=val if val not in ("", None) else None)
            cell.font = f_input()
        row += 1


def fill_vendedor_contacto(ws, vendedor: dict):
    if not vendedor:
        return
    row = CONTACTOS_VENDEDOR_ROW
    if vendedor.get("nombre"):
        ws.cell(row=row, column=2, value=vendedor["nombre"]).font = f_input()
    if vendedor.get("telefono"):
        ws.cell(row=row, column=3, value=vendedor["telefono"]).font = f_input()
    if vendedor.get("email"):
        ws.cell(row=row, column=4, value=vendedor["email"]).font = f_input()
    if vendedor.get("notas"):
        ws.cell(row=row, column=5, value=vendedor["notas"]).font = f_label()


def fill_honorarios(ws, honorarios_servicio):
    if honorarios_servicio in (None, ""):
        return
    cell = ws.cell(row=HONORARIOS_ROW, column=2, value=honorarios_servicio)
    cell.font = f_input()
    cell.fill = fill_yellow


def fill_anuncio(ws, anuncio: dict):
    """Rellena el copy del anuncio (hoja 'Anuncio de venta'). El precio y la
    ficha tecnica de esa hoja son formulas que enlazan a 'Vehiculo y resumen'
    y no se tocan aqui; solo se escribe el texto redactado."""
    if not anuncio:
        return
    mapping = {
        "titulo": ANUNCIO_ROW_TITULO,
        "descripcion_corta": ANUNCIO_ROW_DESC_CORTA,
        "descripcion_larga": ANUNCIO_ROW_DESC_LARGA,
        "hashtags": ANUNCIO_ROW_HASHTAGS,
    }
    for key, row in mapping.items():
        val = anuncio.get(key)
        if val in (None, ""):
            continue
        cell = ws.cell(row=row, column=3, value=val)
        cell.font = f_input()


def main():
    if len(sys.argv) < 3:
        print("Uso: python fill_template.py datos.json /ruta/salida.xlsx [ruta_master.xlsx]")
        sys.exit(1)

    data_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2])
    master_path = Path(sys.argv[3]) if len(sys.argv) > 3 else (
        Path(__file__).resolve().parent.parent / "assets" / "Plantilla_Importacion_Vehiculos_master.xlsx"
    )

    data = json.loads(data_path.read_text(encoding="utf-8"))

    wb = openpyxl.load_workbook(master_path)
    fill_vehicle(wb["Vehiculo y resumen"], data.get("vehicle", {}))
    fill_investigacion(wb["Vehiculo y resumen"], data.get("investigacion", {}))
    fill_comparables(wb[COMPARABLES_SHEET], data.get("comparables", []))
    fill_vendedor_contacto(wb[CONTACTOS_SHEET], data.get("vendedor_contacto", {}))
    fill_honorarios(wb[COMPARABLES_SHEET], data.get("honorarios_servicio"))
    if ANUNCIO_SHEET in wb.sheetnames:
        fill_anuncio(wb[ANUNCIO_SHEET], data.get("anuncio", {}))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"OK: guardado en {out_path}")
    print("Recuerda recalcular con recalc.py (skill xlsx) antes de entregar el archivo.")


if __name__ == "__main__":
    main()
